import os
import re
import pandas as pd
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from tqdm import tqdm
import logging
import sys
import argparse
import unicodedata
import textwrap
import tempfile
import shutil

# ===========================
# 日志配置（默认仅输出到控制台）
# ===========================
def make_logger(verbose: bool = False):
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)]
    )
    return logging.getLogger(__name__)

# 默认 logger（会在 main() 中根据参数重新创建）
logger = make_logger(False)

# ===========================
# 颜色定义
# ===========================
ANSI = {
    'reset': "\033[0m",
    'bold': "\033[1m",
    'cyan': "\033[36m",
    'magenta': "\033[35m",
    'green': "\033[32m",
    'yellow': "\033[33m",
}

AUTHOR = 'zhkali'
REPOS = [
    'https://github.com/ouwenjin/nmap-xml2xlsx',
    'https://gitee.com/zhkali/nmap-xml2xlsx'
]

# 用于去除 ANSI 控制码的正则
_ansi_re = re.compile(r'\x1B\[[0-?]*[ -/]*[@-~]')

def supports_color() -> bool:
    """
    简单检测终端是否支持 ANSI 颜色（Windows 上做了基础兼容判断）
    """
    if sys.platform.startswith('win'):
        return os.getenv('ANSICON') is not None or 'WT_SESSION' in os.environ or sys.stdout.isatty()
    return sys.stdout.isatty()

_COLOR = supports_color()

def strip_ansi(s: str) -> str:
    return _ansi_re.sub('', s)

def visible_width(s: str) -> int:
    s2 = strip_ansi(s)
    w = 0
    for ch in s2:
        if unicodedata.combining(ch):
            continue
        ea = unicodedata.east_asian_width(ch)
        if ea in ('F', 'W'):
            w += 2
        else:
            w += 1
    return w

def pad_visible(s: str, target_visible_len: int) -> str:
    cur = visible_width(s)
    if cur >= target_visible_len:
        return s
    return s + ' ' * (target_visible_len - cur)

def make_lines():
    big_name = r"""
   ███████╗██╗  ██╗██╗  ██╗ █████╗ ██╗      ██╗        
   ╚══███╔╝██║  ██║██║ ██╔╝██╔══██╗██║      ██║        
     ███╔╝ ███████║█████╔╝ ███████║██║      ██║        
    ███╔╝  ██╔══██║██╔═██╗ ██╔══██║██║      ██║        
   ███████╗██║  ██║██║  ██╗██║  ██║███████╗ ██║       
   ╚══════╝╚═╝  ╚═╝╚═╝  ╚═╝╚═╝  ╚═╝╚══════╝ ╚═╝        
"""
    art = textwrap.dedent(big_name)
    art_lines = [ln.rstrip('\n') for ln in art.splitlines() if ln != '']
    author_line = f"作者： {AUTHOR}"
    repo1 = REPOS[0]
    repo2 = REPOS[1]
    return art_lines + [''] + [author_line, repo1, repo2]

def print_banner(use_unicode: bool = True, outer_margin: int = 0, inner_pad: int = 1):
    if use_unicode:
        tl, tr, bl, br, hor, ver = '┌','┐','└','┘','─','│'
    else:
        tl, tr, bl, br, hor, ver = '+','+','+','+','-','|'

    c_reset = ANSI.get('reset','')
    c_bold = ANSI.get('bold','')
    c_cyan = ANSI.get('cyan','')
    c_green = ANSI.get('green','')
    c_yellow = ANSI.get('yellow','')

    raw_lines = make_lines()

    colored = []
    for ln in raw_lines:
        if ln.startswith('作者'):
            colored.append((c_bold + c_green + ln + c_reset) if _COLOR else ln)
        elif ln.startswith('http'):
            colored.append((c_yellow + ln + c_reset) if _COLOR else ln)
        else:
            if ln.strip() == '':
                colored.append(ln)
            else:
                colored.append((c_bold + c_cyan + ln + c_reset) if _COLOR else ln)

    content_max = max((visible_width(x) for x in colored), default=0)
    padded_lines = [pad_visible(ln, content_max) for ln in colored]

    total_inner = inner_pad * 2 + content_max
    width = total_inner + 2

    top = tl + (hor * (width - 2)) + tr
    bottom = bl + (hor * (width - 2)) + br

    pad = ' ' * max(0, outer_margin)

    if _COLOR and use_unicode:
        print(pad + (c_cyan + top + c_reset))
    else:
        print(pad + top)

    left_bar = (c_cyan + ver + c_reset) if _COLOR else ver
    right_bar = (c_cyan + ver + c_reset) if _COLOR else ver
    for pl in padded_lines:
        line_content = (' ' * inner_pad) + pl + (' ' * inner_pad)
        print(pad + left_bar + line_content + right_bar)

    if _COLOR and use_unicode:
        print(pad + (c_cyan + bottom + c_reset))
    else:
        print(pad + bottom)

# ===========================
# 危险端口和服务定义
# ===========================
dangerous_ports = {
    20,21,23,25,53,69,111,110,2049,143,137,135,139,389,445,161,
    512,513,514,873,1433,1521,1529,3306,3389,5000,5432,
    5900,5901,5902,6379,7001,888,9200,9300,11211,27017,27018
}
dangerous_services = {
    'ftp','telnet','smtp','dns','smb','snmp','rsync','oracle','mysql','mysqlx',
    'mariadb','rdp','postgresql','vnc','redis','weblogic_server','elasticsearch',
    'elasticsearch_transport','memcached','mongodb','mongodb_shard_or_secondary',
    'tftp','nfs','pop3','imap','netbios-ns','msrpc','netbios-ssn','ldap',
    'linux rexec','mssql','oracle db','sybase/db2','ilo','any','oracledb',
    'http','linuxrexec','vnc服务'
}

# ===========================
# 校验 IP（返回 bool）
# ===========================
def is_valid_ip(ip: str) -> bool:
    if not ip or str(ip).strip() == '':
        return False
    ip = str(ip).strip()
    ipv4_pattern = r"^(25[0-5]|2[0-4]\d|[01]?\d\d?)" \
                   r"(\.(25[0-5]|2[0-4]\d|[01]?\d\d?)){3}$"
    ipv6_pattern = r"^([0-9a-fA-F]{0,4}:){2,7}[0-9a-fA-F]{0,4}$"
    return re.match(ipv4_pattern, ip) is not None or re.match(ipv6_pattern, ip) is not None

# 从 host 节点提取 IP（优先 ipv4）
def get_ip_from_host(host) -> str:
    # nmap 输出中 address 可能有多个（addrtype="ipv4"/"ipv6"/"mac"）
    addrs = host.findall("address")
    ip = None
    if addrs:
        for a in addrs:
            a_type = a.get("addrtype", "").lower()
            if a_type == "ipv4":
                ip = a.get("addr")
                break
        if not ip:
            # fallback first addr that looks like ip
            for a in addrs:
                candidate = a.get("addr")
                if candidate and is_valid_ip(candidate):
                    ip = candidate
                    break
    # 有些 nmap 会把 address 放在子元素命名空间下，尝试通配
    if not ip:
        # 尝试查找任意属性 addr
        for elem in host.iter():
            if elem.tag.endswith("address") and elem.get("addr"):
                cand = elem.get("addr")
                if is_valid_ip(cand):
                    ip = cand
                    break
    return ip or ""

# ===========================
# 合并所有 Nmap XML 文件
# ===========================
def merge_all_xml(output_file="out.xml"):
    xml_files = [f for f in os.listdir(".") if f.lower().endswith(".xml")]
    if not xml_files:
        logger.warning("没有找到 XML 文件，跳过合并。")
        return None

    logger.info(f"开始合并 {len(xml_files)} 个 XML 文件 -> {output_file}")
    try:
        main_tree = ET.parse(xml_files[0])
        main_root = main_tree.getroot()
    except Exception as e:
        logger.error(f"解析第一个 XML ({xml_files[0]}) 失败: {e}")
        return None

    for xml_file in xml_files[1:]:
        try:
            tree = ET.parse(xml_file)
            root = tree.getroot()
            # 找到所有 host 节点并追加
            for host in root.findall(".//host"):
                main_root.append(host)
        except Exception as e:
            logger.error(f"合并文件 {xml_file} 出错: {e}")

    try:
        main_tree.write(output_file, encoding="utf-8", xml_declaration=True)
        logger.info(f"XML 合并完成，结果保存为 {output_file}")
        return output_file
    except Exception as e:
        logger.error(f"保存合并 XML 失败: {e}")
        return None

# ===========================
# 解析 Nmap XML
# ===========================
def parse_nmap_xml(xml_file):
    results = []
    if not os.path.exists(xml_file):
        logger.warning(f"文件不存在: {xml_file}")
        return results
    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()
        hosts = root.findall(".//host")
        for h_index, host in enumerate(tqdm(hosts, desc=f"解析Nmap: {xml_file}", unit="host")):
            ip = get_ip_from_host(host)
            if not is_valid_ip(ip):
                logger.debug(f"[Nmap] 文件 {xml_file} Host#{h_index} IP 无效或缺失: {ip}")

            # port 元素可能在 host/ports/port 下
            for port in host.findall(".//port"):
                proto = port.get("protocol") or ""
                portid = port.get("portid") or ""
                # state
                state_elem = port.find("state")
                state = state_elem.get("state") if state_elem is not None and state_elem.get("state") else ""
                # service
                service_elem = port.find("service")
                service = service_elem.get("name") if service_elem is not None and service_elem.get("name") else ""
                results.append({
                    "IP": ip,
                    "端口/协议": f"{portid}/{proto}" if portid else f"/{proto}",
                    "状态": state,
                    "服务": service,
                    "端口用途": "",
                })
    except Exception as e:
        logger.error(f"解析 Nmap 文件 {xml_file} 出错: {e}")
    return results

# ===========================
# 解析 Excel/CSV 表格（增加编码回退和列名模糊匹配）
# ===========================
def parse_table(file_path):
    results = []
    if not os.path.exists(file_path):
        logger.error(f"文件不存在: {file_path}")
        return results
    try:
        # 自动尝试编码
        if file_path.lower().endswith(".xlsx") or file_path.lower().endswith(".xls"):
            df = pd.read_excel(file_path, engine="openpyxl")
        else:
            try:
                df = pd.read_csv(file_path, encoding="utf-8")
            except Exception:
                df = pd.read_csv(file_path, encoding="gbk", errors="ignore")
        if df is None or df.empty:
            logger.warning(f"文件为空: {file_path}")
            return results

        # 列映射（保留原逻辑，但做小写匹配与模糊匹配）
        col_map = {
            "IP": ["ip","地址","host"],
            "端口/协议": ["端口/协议","端口","port"],
            "状态": ["状态","state","开放状态"],
            "服务": ["服务","service","协议"],
            "端口用途": ["端口用途","用途","备注","remark"]
        }
        cols_lower = {c.lower(): c for c in df.columns}
        real_cols = {}
        for std_col, aliases in col_map.items():
            found = None
            for alias in aliases:
                if alias in df.columns:
                    found = alias
                    break
                if alias.lower() in cols_lower:
                    found = cols_lower[alias.lower()]
                    break
            # 如果还没找到，尝试模糊匹配（包含关键字）
            if not found:
                for actual in df.columns:
                    al = actual.lower()
                    for alias in aliases:
                        if alias in al or alias.lower() in al:
                            found = actual
                            break
                    if found:
                        break
            real_cols[std_col] = found

        for i, row in tqdm(df.iterrows(), total=len(df), desc=f"解析表格: {file_path}", unit="行"):
            raw_ip = row.get(real_cols["IP"], "") if real_cols["IP"] else ""
            ip = str(raw_ip).strip() if pd.notna(raw_ip) else ""
            if ip and not is_valid_ip(ip):
                logger.debug(f"[表格] 文件 {file_path} 行 {i+2} IP 看起来无效: {ip}")

            port_proto = row.get(real_cols["端口/协议"], "") if real_cols["端口/协议"] else ""
            port_proto = "" if pd.isna(port_proto) else str(port_proto).strip()
            if port_proto and "/" not in str(port_proto):
                port_proto = f"{port_proto}/tcp"

            state = row.get(real_cols["状态"], "") if real_cols["状态"] else ""
            state = "" if pd.isna(state) else str(state).strip()
            service = row.get(real_cols["服务"], "") if real_cols["服务"] else ""
            service = "" if pd.isna(service) else str(service).strip()
            remark = row.get(real_cols["端口用途"], "") if real_cols["端口用途"] else ""
            remark = "" if pd.isna(remark) else str(remark).strip()

            results.append({
                "IP": ip,
                "端口/协议": port_proto,
                "状态": state,
                "服务": service,
                "端口用途": remark,
            })
    except Exception as e:
        logger.error(f"解析文件 {file_path} 出错: {e}")
    return results

# ===========================
# 标记危险端口/服务（更宽容处理）
# ===========================
def mark_dangerous(df):
    def check(row):
        port = None
        try:
            port_str = str(row.get("端口/协议", "")).split("/")[0]
            if port_str and re.match(r"^\d+$", port_str):
                port = int(port_str)
        except Exception:
            port = None
        service = str(row.get("服务", "")).strip().lower()
        # check service tokens
        service_tokens = set(re.split(r'[\s/_\-]+', service))
        if (port in dangerous_ports) or (service and (service in dangerous_services or bool(service_tokens & dangerous_services))):
            return "危险端口不允许对外开放"
        return ""
    df["是否必要开放"] = df.apply(check, axis=1)
    return df

# ===========================
# Excel 美化（冻结表头、自动筛选、表头样式）
# ===========================
def format_excel(file_path):
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        font = Font(name="宋体", size=12)
        bold_font = Font(name="宋体", size=12, bold=True)
        red_font = Font(name="宋体", size=12, color="FFFF0000")
        header_fill = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")
        align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        column_widths = {"A":36,"B":12,"C":12,"D":18,"E":11,"F":28}
        # 基于现有列数设置（避免列名短于映射）
        for i, col in enumerate(ws.iter_cols(1, ws.max_column), start=1):
            col_letter = col[0].column_letter
            width = column_widths.get(col_letter, 18)
            ws.column_dimensions[col_letter].width = width

        # header 行格式化
        for cell in ws[1]:
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = align

        # 其他行基础字体
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = font
                cell.alignment = Alignment(vertical="top", wrap_text=True)

                if cell.value == "危险端口不允许对外开放":
                    cell.font = red_font

        # 冻结首行
        ws.freeze_panes = "A2"

        # 自动筛选（覆盖整个表）
        try:
            ws.auto_filter.ref = ws.dimensions
        except Exception:
            pass

        wb.save(file_path)
    except Exception as e:
        logger.error(f"格式化 Excel 失败: {e}")

# ===========================
# 去重逻辑（归一化后去重）
# ===========================
def auto_dedup(df):
    if df.empty:
        return df, "none"
    before = len(df)
    # 先归一化关键字段
    for col in ["IP","端口/协议","服务","状态","端口用途"]:
        if col in df.columns:
            df[col] = df[col].astype(str).fillna("").map(lambda x: re.sub(r'\s+', ' ', x).strip().lower())
    df.drop_duplicates(subset=["IP","端口/协议","服务","状态","端口用途"], inplace=True)
    after = len(df)
    mode = f"strict ({before-after} 行被删除)"
    return df, mode

# ===========================
# 主函数
# ===========================
def main():
    global logger, _COLOR
    parser = argparse.ArgumentParser(description='合并 Nmap XML 和 Excel/CSV 扫描结果，生成端口调研表，并打印作者横幅')
    parser.add_argument('--no-unicode', dest='no_unicode', action='store_true',
                        help='强制使用 ASCII 框（不使用 Unicode 盒绘字符）')
    parser.add_argument('--margin', type=int, default=0, help='横幅左侧外边距空格数（默认 0）')
    parser.add_argument('--pad', type=int, default=1, help='横幅内部左右边距（默认 1）')
    parser.add_argument('--input', '-i', default="开放端口.xlsx", help='输入 Excel/CSV 文件路径（默认 开放端口.xlsx）')
    parser.add_argument('--output', '-o', default="端口调研表.xlsx", help='输出文件名（默认 端口调研表.xlsx）')
    parser.add_argument('--temp-xml', default="out.xml", help='临时合并的 XML 文件名（默认 out.xml）')
    parser.add_argument('--cleanup', action='store_true', help='处理完成后删除临时 out.xml')
    parser.add_argument('--no-color', action='store_true', help='禁用颜色输出')
    parser.add_argument('--verbose', action='store_true', help='开启详细日志(DEBUG)')
    args = parser.parse_args()

    # 重新创建 logger（使用 --verbose）
    logger = make_logger(args.verbose)
    if args.no_color:
        _COLOR = False

    print_banner(use_unicode=not args.no_unicode, outer_margin=args.margin, inner_pad=max(0, args.pad))

    all_results = []

    # 第一步：合并 XML（如果目录下存在 .xml 文件）
    merged_xml = merge_all_xml(args.temp_xml)

    # 第二步：解析 Excel/CSV（支持指定输入文件）
    input_file = args.input
    logger.info(f"解析表格文件: {input_file}")
    all_results.extend(parse_table(input_file))

    # 第三步：解析 out.xml
    if merged_xml:
        all_results.extend(parse_nmap_xml(merged_xml))

    if not all_results:
        logger.error("未找到可解析数据。")
        return

    df = pd.DataFrame(all_results)

    # 去重
    df, mode = auto_dedup(df)
    logger.info(f"自动去重模式：{mode}，最终 {len(df)} 行")

    # 标注危险
    df = mark_dangerous(df)

    output_file = args.output
    # 原子写入：先写临时文件再替换
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(tmp_fd)
    try:
        df.to_excel(tmp_path, index=False)
        format_excel(tmp_path)
        # 替换目标文件（原子）
        shutil.move(tmp_path, output_file)
        logger.info(f"处理完成，结果保存为 {output_file}")
    except Exception as e:
        logger.error(f"保存输出文件失败: {e}")
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
    # cleanup 临时 xml
    if args.cleanup and merged_xml and os.path.exists(merged_xml):
        try:
            os.remove(merged_xml)
            logger.info(f"已删除临时文件: {merged_xml}")
        except Exception as e:
            logger.debug(f"删除临时文件失败: {e}")

if __name__ == "__main__":
    main()
